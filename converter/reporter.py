"""Gera os entregáveis finais: xlsx por estrutura, relatório MD e zip."""

import io
import json
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .schemas import get_structure
from .validator import ValidationResult


HANDIT_NAVY = "1B355B"
HANDIT_GREEN = "00C389"


def _df_to_xlsx_bytes(df: pd.DataFrame, sheet_title: str = "Layout de Carga") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial", color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HANDIT_NAVY)

    # Data
    for _, row in df.iterrows():
        ws.append([str(v) if pd.notna(v) else "" for v in row.tolist()])

    # Font padrão e largura
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="Arial")
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(12, min(40, len(str(col_name)) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_report_md(
    client_name: str,
    source_filename: str,
    file_kind: str,
    mapping_or_extraction: Dict[str, Any],
    dfs: Dict[str, pd.DataFrame],
    validations: Dict[str, ValidationResult],
) -> str:
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    parts = [
        f"# Relatório de Conversão - {client_name}",
        "",
        f"**Arquivo fonte:** `{source_filename}`",
        f"**Tipo detectado:** {file_kind}",
        f"**Gerado em:** {now}",
        "",
        "---",
        "",
        "## 1. Arquivos gerados",
        "",
        "| Estrutura | Linhas | Erros | Avisos |",
        "|---|---:|---:|---:|",
    ]
    for sid, df in dfs.items():
        s = get_structure(sid)
        v = validations.get(sid)
        parts.append(
            f"| {s.label} | {len(df)} | "
            f"{len(v.errors) if v else 0} | {len(v.warnings) if v else 0} |"
        )

    parts += ["", "---", "", "## 2. Validações"]
    for sid, v in validations.items():
        s = get_structure(sid)
        parts += ["", f"### {s.label}", ""]
        if not v.errors and not v.warnings:
            parts.append("Nenhum erro ou alerta.")
        if v.errors:
            parts.append("**Erros:**")
            parts += [f"- {e}" for e in v.errors]
        if v.warnings:
            parts.append("")
            parts.append("**Avisos:**")
            parts += [f"- {w}" for w in v.warnings]
        if v.metrics:
            parts += ["", "**Métricas:**", "```json", json.dumps(v.metrics, indent=2, default=str), "```"]

    parts += ["", "---", "", "## 3. Decisões e mapeamento usado", "", "```json"]
    parts.append(json.dumps(mapping_or_extraction, indent=2, ensure_ascii=False)[:8000])
    parts.append("```")

    parts += [
        "",
        "---",
        "",
        "## 4. Checklist antes do upload no FP&A Base",
        "",
        "- [ ] Revisar EMPRESA_COD e EMPRESA_DESC na Estrutura Empresarial",
        "- [ ] Conferir CCs e contas órfãs no razão",
        "- [ ] Validar totais por filial com extrato do cliente",
        "- [ ] Decidir estratégia para lançamentos sem CC (atribuir 0 ou reclassificar)",
        "",
    ]
    return "\n".join(parts)


def generate_outputs(
    client_name: str,
    source_filename: str,
    file_kind: str,
    mapping_or_extraction: Dict[str, Any],
    dfs: Dict[str, pd.DataFrame],
    validations: Dict[str, ValidationResult],
) -> bytes:
    """Gera um zip em memória com xlsx por estrutura + relatório MD."""
    buf = io.BytesIO()
    safe_client = "".join(c if c.isalnum() else "_" for c in client_name) or "cliente"

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # xlsx por estrutura
        structure_order = [
            "estrutura_empresarial",
            "centro_de_custo",
            "plano_de_contas",
            "razao_contabil",
        ]
        idx_to_prefix = {s: f"{i+1:02d}" for i, s in enumerate(structure_order)}
        for sid, df in dfs.items():
            prefix = idx_to_prefix.get(sid, "99")
            struct = get_structure(sid)
            filename = f"{prefix}_{struct.label.replace(' ', '')}_{safe_client}.xlsx"
            zf.writestr(filename, _df_to_xlsx_bytes(df))

        # Relatório
        report = _render_report_md(
            client_name=client_name,
            source_filename=source_filename,
            file_kind=file_kind,
            mapping_or_extraction=mapping_or_extraction,
            dfs=dfs,
            validations=validations,
        )
        zf.writestr(f"Relatorio_Conversao_{safe_client}.md", report)

    return buf.getvalue()
